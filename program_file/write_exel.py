import pandas as pd
import openpyxl
import os
from openpyxl import load_workbook, Workbook

# 解析済みのファイルを取得
analysis_directory_path = './output'
exclude_folders = {'output_log'}  # 除外するフォルダのセット

# 特定のファイルを再帰的に取得していく関数
def get_specific_files_recursively(target_file_name):
    get_files = []
    for root, dirs, files in os.walk(analysis_directory_path):
        # 除外するフォルダをスキップ
        dirs[:] = [d for d in dirs if d not in exclude_folders]
        # ファイルをリスト内に取得
        for file in files:
            if file == target_file_name:
                get_files.append(os.path.join(root, file))
    # ファイル名でソート(番号順に処理)
    get_files.sort()
    return get_files
    
# エクセルファイルを再帰的に取得
excel_files = get_specific_files_recursively('analysis.xlsx')

if not excel_files:
    print("エラー：参照するエクセルファイルが見つかりませんでした。")
else:
    # 新しいExcelワークブックを作成
    output_excel_path = './output/analysis_all_data.xlsx'
    workbook = Workbook()
    
    # 各ファイルを処理
    for file_path in excel_files:
        # ファイル名をシート名として使用
        sheet_name = os.path.basename(os.path.dirname(file_path))
        sheet = workbook.create_sheet(title=sheet_name)
        
        # ファイルの内容を読み込む
        source_workbook = load_workbook(file_path, data_only=False)
        source_sheet = source_workbook.active
        
        # データを新しいシートに書き込む
        for row in source_sheet.iter_rows():
            new_row = [cell.value for cell in row]
            sheet.append(new_row)
    
    # デフォルトのシートを削除
    if 'Sheet' in workbook.sheetnames:
        workbook.remove(workbook['Sheet'])

    # 変更をエクセルファイルに保存する
    workbook.save(output_excel_path)
    print(f"データが {output_excel_path} に書き込まれました。")