import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
import os

# 解析済みのファイルを取得
analysis_directory_path = './output'
# 除外するフォルダのセット
exclude_folders = {'output_log'}

# 特定のファイルを再帰的に取得（例：.xlsxファイル）
excel_files = []
for root, dirs, files in os.walk(analysis_directory_path):
    # 除外するフォルダをスキップ
    dirs[:] = [d for d in dirs if d not in exclude_folders]

    for file in files:
        if file == 'photo.xlsx':
            excel_files.append(os.path.join(root, file))

# 新しいExcelワークブックを作成
output_excel_path = './output/photo_data.xlsx'
with pd.ExcelWriter(output_excel_path, engine='openpyxl') as writer:
    # 各ファイルを処理
    for file_path in excel_files:
        # ファイル名をシート名として使用
        sheet_name = os.path.basename(os.path.dirname(file_path))
    
        # ファイルの内容を読み込む
        df = pd.read_excel(file_path)
        
        # データフレームの内容を新しいシートに書き込む
        df.to_excel(writer, sheet_name=sheet_name, index=False)

# openpyxlで新しいExcelファイルを開く
wb = openpyxl.load_workbook(output_excel_path)
black_fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')

# 各シートを処理
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    # データフレーム内の "●" の位置を特定して塗りつぶし、"〇" を空白に置換
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == "●":
                cell.fill = black_fill
            elif cell.value == "〇":
                cell.value = ""

# ファイルを保存
wb.save(output_excel_path)

print(f"データが {output_excel_path} に書き込まれました。")