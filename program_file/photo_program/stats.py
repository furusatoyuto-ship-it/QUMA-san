import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
import os

# 解析済みのファイルを取得
analysis_directory_path = './output'

# ディレクトリ内のすべてのフォルダーを取得
folders = [f for f in os.listdir(analysis_directory_path) if os.path.isdir(os.path.join(analysis_directory_path, f))]

# すべてのフォルダーでループ
for folder in folders:
    if not folder == "output_log":
        # 解析するファイルパスを指定
        file_path = os.path.join(analysis_directory_path, folder, 'analysis.xlsx')
        # ヘッダーなしで読み込む
        df = pd.read_excel(file_path, header=None)

        # "excluded"をカウントする変数
        excluded_count = -1

        # 新しいデータフレームを作成
        new_df = pd.DataFrame()

        # 行ごとにループ
        for index, row in df.iterrows():
            # 初期のスコアを設定
            row_score = 0
            cell_no = 0
            me_mass = 0
            unme_mass = 0
            total_score = 0
            ans_score = 0

            # 最初のセルを取得
            first_cell = str(row[0])

            # 処理をスキップする条件
            excluded_items = ["excluded", "位置1"]
            if first_cell in excluded_items:
                excluded_count += 1
            else:
                # 列ごとにループ(実質セルごと)
                for col_index, value in enumerate(row):
                    # セルの値を文字列に変換
                    cell_values = str(value)

                    # セルごとにループ
                    for cell_value in cell_values:
                        # cell_noに+1
                        cell_no += 1
                        
                        # ●と○でスコアを産出
                        if cell_value == "●":
                            # 二進数の様にカウント一列目は2^0
                            row_score += 2 ** cell_no
                            me_mass += 1
                        else:
                            row_score += 0
                            unme_mass += 1
                    
                    total_mass = me_mass + unme_mass

                    # スコアに質量の影響を追加
                    total_score = 2 ** total_mass * me_mass
                    ans_score = row_score + total_score
                
                # 行の最後にスコアを追加
                new_row = row.tolist() + [ans_score]
                new_df = pd.concat([new_df, pd.DataFrame([new_row])], ignore_index=True)

        if new_df.empty:
            print(folder, "のデータフレームは空です")
        else:
            # スコアをもとにデータフレームを並べ替え
            sorted_df = new_df.sort_values(by=new_df.columns[-1], ascending=True).reset_index(drop=True)

            # 並べ替えたデータフレームを表示
            print(sorted_df)
            print("excludedの数", excluded_count)

            # データフレーム全体で "○" を空白に置換
            sorted_df = sorted_df.replace("○", "")

            output_file_path = os.path.join(analysis_directory_path, folder, 'photo.xlsx')
            sorted_df.to_excel(output_file_path, index=False, engine='openpyxl')

            # openpyxlでExcelファイルを開く
            wb = openpyxl.load_workbook(output_file_path)
            ws = wb.active

            # 黒で塗りつぶすためのスタイルを定義
            black_fill = openpyxl.styles.PatternFill(start_color='000000', end_color='000000', fill_type='solid')

            # データフレーム内の "●" の位置を特定して塗りつぶす
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value == "●":
                        cell.fill = black_fill

            # ファイルを保存
            wb.save(output_file_path)