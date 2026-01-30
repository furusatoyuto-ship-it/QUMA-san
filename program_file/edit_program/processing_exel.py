"""
エクセルのデータを加工するプログラム
"""
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

#新しい列に関数を挿入する
def add_function(sheet, value_name):
    # 新しい列を挿入する位置を指定
    new_column_idx = sheet.max_column + 1
    # 新しい列にExcelの関数を追加する
    for row in range(2, sheet.max_row + 1):  # 1行目はヘッダーと仮定してスキップ
        last_column_letter = get_column_letter(new_column_idx - 1)
        cell = sheet.cell(row=row, column=new_column_idx)
        #value_nameに対して関数を指定し、挿入する
        value_function = {
            'Methylation_level': f'=COUNTIF(A{row}:{last_column_letter}{row}, "●")',
            'percentage': f'={last_column_letter}{row}/{sheet.max_column - 2}'
            }
        cell.value = value_function[value_name]

    # 新しい列のヘッダーを設定
    sheet.cell(row=1, column=new_column_idx, value=value_name)

# 特定のセルに関数を挿入する
def add_graph(sheet, add_name, add_column):
    #入れたい列記号（ex.A,B)を取得
    column_letter = get_column_letter(add_column)
    #入れたい行番号を指定
    row_number = 2
    #名前を２行目に挿入
    sheet[f'{column_letter}{row_number}'] = add_name
    #名前に応じた関数を３行目に挿入
    add_function = {
        "all": '=COUNTA(A:A) - 1',
        "used": '=COUNTA(A:A) - 1 - COUNTIF(A:A,"excluded")' ,
        "excluded": '=COUNTIF(A:A,"excluded")'
        }
    sheet[f'{column_letter}{row_number+1}'] = add_function[add_name]

# 解析済みのファイルを取得
analysis_directory_path = './output'
exclude_folders = {'output_log'}  # 除外するフォルダのセット

# 特定のファイルを再帰的に取得していく関数
def get_specific_files_recursively(target_file_name):
    get_files = []
    for root, dirs, files in os.walk(analysis_directory_path):
        # 除外するフォルダをスキップ
        dirs[:] = [d for d in dirs if d not in exclude_folders]
        # ファイルをリスト内に取得
        for file in files:
            if file == target_file_name:
                get_files.append(os.path.join(root, file))
    # ファイル名でソート(番号順に処理)
    get_files.sort()
    return get_files

# エクセルファイルを再帰的に取得する
xlsx_files = get_specific_files_recursively('analysis.xlsx')

if not xlsx_files:
    print("⚠ analysis.xlsx ファイルが見つかりませんでした。")
    exit()
else:
    for file_path in xlsx_files:
        # エクセルファイルを読み込む
        workbook = load_workbook(file_path)
        sheet = workbook.active

        #メチル化の数を調べる列を挿入
        add_function(sheet, 'Methylation_level')

        #メチル化割合を調べる列を挿入
        add_function(sheet, 'percentage')

        #現時点で一番最後の列番号を指定
        last_column = sheet.max_column

        #使用した実験の件数を最後の列から２列横に挿入
        add_graph(sheet, "all", last_column+2)

        #使用できる実験の件数を最後の列から３列横に挿入
        add_graph(sheet, "used", last_column+3)

        #excludedの件数を最後の列から４列横に挿入
        add_graph(sheet, "excluded", last_column+4)

        # 変更をエクセルファイルに保存する
        workbook.save(file_path)
        print(f"{file_path}に保存しました")